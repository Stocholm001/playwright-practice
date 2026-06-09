# import json
# from openpyxl import load_workbook

# EXCEL_PATH = r'D:\TestResults\SauceDemo_TestCases.xlsx'
# RESULTS_PATH = r'D:\TestResults\results.json'

# wb = load_workbook(EXCEL_PATH)
# ws = wb['Authentication']

# with open(RESULTS_PATH, 'r', encoding='utf-8') as f:
#     results = json.load(f)

# for row in ws.iter_rows(min_row=4):
#     tc_id = row[0].value
#     if tc_id in results:
#         row[5].value = results[tc_id]  # column F = Status

# wb.save(EXCEL_PATH)
# print(f"Updated: {EXCEL_PATH}")



import re
import json
import os
import glob
from datetime import datetime
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
EXCEL_PATH = os.path.join(BASE_DIR, 'TestCase', 'SauceDemo_TestCases.xlsx')
RESULTS_PATH = os.path.join(BASE_DIR, 'TestCase', 'results.json')

def clean_text(text):
    text = text.replace('\n', ' | ')
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', text)
    return text

# 1. หา source — Result ล่าสุด หรือ template ถ้ายังไม่มี
result_files = sorted(glob.glob(os.path.join(BASE_DIR, 'TestCase', 'Result_*.xlsx')))
SOURCE_PATH = result_files[-1] if result_files else EXCEL_PATH

# 2. โหลด source
wb = load_workbook(SOURCE_PATH)
sheet_name = next(s for s in wb.sheetnames if 'Authentication' in s)
ws = wb[sheet_name]

# 3. โหลด results.json
with open(RESULTS_PATH, 'r', encoding='utf-8') as f:
    results = json.load(f)

# 4. อัปเดตเฉพาะ TC ที่รัน — ที่เหลือคงค่าเดิม
for row in ws.iter_rows(min_row=4):
    tc_id = row[0].value
    if tc_id in results:
        data = json.loads(results[tc_id])
        row[5].value = data['status']
        row[7].value = clean_text(data['note'])

# 5. ลบ Result เก่า แล้ว save ใหม่
for f in result_files:
    os.remove(f)

timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
output_path = os.path.join(BASE_DIR, 'TestCase', f'Result_{timestamp}.xlsx')
wb.save(output_path)

print(f"Updated: {output_path}")